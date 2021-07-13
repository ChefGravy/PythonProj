# -*- coding: utf-8 -*-
"""
Created on Tue Sep 22 20:05:45 2020

@author: Ruslan Polyak
"""
# -*- coding: utf-8 -*-


from tkinter import Label, Tk, Entry, IntVar, W, Button, filedialog
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import math
import tkinter as tk
import pandas as pd
import sys    

def execute_command():    
    root = tk.Tk()
    root.withdraw()
    
    # ask for file location
    file_loc = filedialog.askopenfilename()
    
    # load the work
    workbook = load_workbook(file_loc, data_only=True)
    # change the default working directory to folder path
    file_path = file_loc.replace("/","\\")
    
    
    # change working directory
    os.chdir(os.path.dirname(file_path))
    try:
        tk.messagebox.showinfo("Working Directory", "Your file will be dropped here: {}. Look for a file with prefix 'finalproduct'".format(os.path.dirname(file_path)))
    except:
        print("Your file will be dropped here: {}. Look for a file with prefix 'finalproduct'".format(os.path.dirname(file_path)))
    
    # this will give file name. 
    # for extension, use [-4:], but not needed
    filepath_split = file_path.split("\\")
    
    # create filename
    file_name = (filepath_split[len(filepath_split)-1:len(filepath_split)][0][:-5])
    file_ext = filepath_split[len(filepath_split)-1:len(filepath_split)][0][-4:]
    filename = "finalproduct_{}.{}".format(file_name, 'xlsx')
    
    # data structure dic to store values
    final_values = {}
    try: 
        # loop through all the sheets and gather data in dic
        for value in workbook.sheetnames[tab_one.get():len(workbook.sheetnames)]:
            ws = workbook[value]
            for val_cost, val_bene in zip(ws.iter_rows(min_row=c_row.get(), max_row=c_row.get(), min_col=f_col.get(),max_col=l_col.get(), values_only=True),ws.iter_rows(min_row=b_row.get(), max_row=b_row.get(), min_col=f_col.get(),max_col=l_col.get(), values_only=True)):
                # manually add cells for financial metrics
                final_values[('{}'.format(value), 'cost')] = val_cost
                final_values[('{}'.format(value), 'benefit')] = val_bene                 
                final_values[('{}'.format(value), 'B/C')] = [ws['B45'].value]
                final_values[('{}'.format(value), 'MIRR')] = [ws['B46'].value]
                final_values[('{}'.format(value), 'Payback')] = [(ws['B47'].value)]
                final_values[('{}'.format(value), 'NPV')] = [(ws['B48'].value)]
                
        df = pd.DataFrame.from_dict(final_values, orient='index')
        df.to_excel(filename)
        try:
            tk.messagebox.showinfo("Success", "Consolidation is complete!")
        except:
            print("Consolidation is complete!")
    except Exception as e:
         print(e)

window = Tk()

window.wm_title("Consolidator")

l1 = Label(window, width=25, text="First Column", anchor=W)
l1.grid(row=0,column=0)

l2 = Label(window, width=25, text="Last Column", anchor=W)
l2.grid(row=1,column=0)

l3 = Label(window, width=25, text="Cost Row", anchor=W)
l3.grid(row=2,column=0)

l4 = Label(window, width=25, text="Benefit Row", anchor=W)
l4.grid(row=3,column=0)

l5 = Label(window, width=25, text="First Tab", anchor=W)
l5.grid(row=4,column=0)

#datatype for window entry
f_col=IntVar()
e1=Entry(window, width=20, textvariable=f_col)
e1.grid(row=0,column=1)

l_col=IntVar()
e2=Entry(window, width=20, textvariable=l_col)
e2.grid(row=1,column=1)

c_row=IntVar()
e3=Entry(window, width=20, textvariable=c_row)
e3.grid(row=2,column=1)

b_row=IntVar()
e4=Entry(window, width=20, textvariable=b_row)
e4.grid(row=3,column=1)

tab_one=IntVar()
e5=Entry(window, width=20, textvariable=tab_one)
e5.grid(row=4,column=1)

b1 = Button(window, text="Execute", width=12, command = execute_command)
b1.grid(row=0,column=2)

b2 = Button(window, text="Close", width=12, command = window.destroy)
b2.grid(row=1,column=2)

window.mainloop()


